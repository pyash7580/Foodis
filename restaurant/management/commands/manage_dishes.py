from django.core.management.base import BaseCommand
from client.models import MenuItem, Restaurant, Category
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
import os


class Command(BaseCommand):
    help = 'Manage dishes via Excel: export, import, or create template'

    def add_arguments(self, parser):
        parser.add_argument(
            '--action',
            type=str,
            required=True,
            choices=['export', 'import', 'template'],
            help='Action to perform: export, import, or template'
        )
        parser.add_argument(
            '--file',
            type=str,
            default=None,
            help='Excel file path (for import/export)'
        )
        parser.add_argument(
            '--restaurant-id',
            type=int,
            default=None,
            help='Filter dishes by restaurant ID (export only)'
        )

    def handle(self, *args, **options):
        action = options['action']
        file_path = options['file']
        restaurant_id = options['restaurant_id']

        if action == 'export':
            self.export_dishes(file_path, restaurant_id)
        elif action == 'import':
            self.import_dishes(file_path)
        elif action == 'template':
            self.create_template(file_path)

    def export_dishes(self, file_path, restaurant_id=None):
        """Export all dishes to Excel file."""
        if not file_path:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            file_path = f'dishes_export_{timestamp}.xlsx'

        self.stdout.write(self.style.SUCCESS(f'Exporting dishes to {file_path}...'))

        # Query dishes
        queryset = MenuItem.objects.select_related('restaurant', 'category').all()
        if restaurant_id:
            queryset = queryset.filter(restaurant_id=restaurant_id)

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Dishes"

        # Header styling
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        # Headers
        headers = [
            'ID', 'Name', 'Description', 'Price', 'Restaurant ID', 'Restaurant Name',
            'Category ID', 'Category Name', 'Veg Type', 'Is Available', 'Rating',
            'Total Orders', 'Has Image', 'Created At'
        ]

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Data rows
        for row_num, dish in enumerate(queryset, 2):
            ws.cell(row=row_num, column=1).value = dish.id
            ws.cell(row=row_num, column=2).value = dish.name
            ws.cell(row=row_num, column=3).value = dish.description or ''
            ws.cell(row=row_num, column=4).value = float(dish.price)
            ws.cell(row=row_num, column=5).value = dish.restaurant.id if dish.restaurant else ''
            ws.cell(row=row_num, column=6).value = dish.restaurant.name if dish.restaurant else ''
            ws.cell(row=row_num, column=7).value = dish.category.id if dish.category else ''
            ws.cell(row=row_num, column=8).value = dish.category.name if dish.category else ''
            ws.cell(row=row_num, column=9).value = dish.veg_type
            ws.cell(row=row_num, column=10).value = 'Yes' if dish.is_available else 'No'
            ws.cell(row=row_num, column=11).value = float(dish.rating) if dish.rating else 0
            ws.cell(row=row_num, column=12).value = dish.total_orders
            ws.cell(row=row_num, column=13).value = 'Yes' if dish.image else 'No'
            ws.cell(row=row_num, column=14).value = dish.created_at.strftime('%Y-%m-%d %H:%M:%S') if dish.created_at else ''

        # Adjust column widths
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

        # Save workbook
        wb.save(file_path)
        self.stdout.write(self.style.SUCCESS(
            f'[OK] Exported {queryset.count()} dishes to {file_path}'
        ))

    def import_dishes(self, file_path):
        """Import dishes from Excel file."""
        if not file_path or not os.path.exists(file_path):
            self.stdout.write(self.style.ERROR('Error: File path required and must exist'))
            return

        self.stdout.write(self.style.SUCCESS(f'Importing dishes from {file_path}...'))

        try:
            wb = load_workbook(file_path)
            ws = wb.active

            created_count = 0
            updated_count = 0
            error_count = 0
            errors = []

            # Skip header row
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
                try:
                    dish_id, name, description, price, restaurant_id, _, category_id, _, veg_type, is_available, *rest = row

                    # Validate required fields
                    if not name or not price or not restaurant_id:
                        errors.append(f'Row {row_num}: Missing required fields (name, price, or restaurant_id)')
                        error_count += 1
                        continue

                    # Get or validate restaurant
                    try:
                        restaurant = Restaurant.objects.get(id=restaurant_id)
                    except Restaurant.DoesNotExist:
                        errors.append(f'Row {row_num}: Restaurant ID {restaurant_id} not found')
                        error_count += 1
                        continue

                    # Get or create category
                    category = None
                    if category_id:
                        try:
                            category = Category.objects.get(id=category_id)
                        except Category.DoesNotExist:
                            errors.append(f'Row {row_num}: Category ID {category_id} not found (skipping category)')

                    # Parse is_available
                    if isinstance(is_available, str):
                        is_available = is_available.lower() in ['yes', 'true', '1']
                    elif is_available is None:
                        is_available = True

                    # Create or update dish
                    defaults = {
                        'name': name,
                        'description': description or '',
                        'price': float(price),
                        'restaurant': restaurant,
                        'category': category,
                        'veg_type': veg_type or 'VEG',
                        'is_available': is_available,
                    }

                    if dish_id:
                        # Update existing dish
                        dish, created = MenuItem.objects.update_or_create(
                            id=dish_id,
                            defaults=defaults
                        )
                        if created:
                            created_count += 1
                        else:
                            updated_count += 1
                    else:
                        # Create new dish
                        MenuItem.objects.create(**defaults)
                        created_count += 1

                except Exception as e:
                    errors.append(f'Row {row_num}: {str(e)}')
                    error_count += 1

            # Summary
            self.stdout.write(self.style.SUCCESS('\n=== Import Summary ==='))
            self.stdout.write(f'Created: {created_count}')
            self.stdout.write(f'Updated: {updated_count}')
            self.stdout.write(f'Errors: {error_count}')

            if errors:
                self.stdout.write(self.style.WARNING('\n=== Errors ==='))
                for error in errors[:10]:  # Show first 10 errors
                    self.stdout.write(self.style.ERROR(error))
                if len(errors) > 10:
                    self.stdout.write(f'... and {len(errors) - 10} more errors')

        except Exception as e:
            self.stdout.write(self.style.ERROR(f'Critical error: {str(e)}'))

    def create_template(self, file_path):
        """Create a template Excel file for importing dishes."""
        if not file_path:
            file_path = 'dishes_template.xlsx'

        self.stdout.write(self.style.SUCCESS(f'Creating template file: {file_path}...'))

        wb = Workbook()
        ws = wb.active
        ws.title = "Dishes Template"

        # Header styling
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        # Headers (ID optional for new dishes)
        headers = [
            'ID (Optional)', 'Name*', 'Description', 'Price*', 'Restaurant ID*',
            'Restaurant Name (Info)', 'Category ID', 'Category Name (Info)',
            'Veg Type', 'Is Available'
        ]

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Example row
        ws.cell(row=2, column=1).value = ''  # ID (leave empty for new)
        ws.cell(row=2, column=2).value = 'Paneer Butter Masala'
        ws.cell(row=2, column=3).value = 'Delicious paneer in rich tomato gravy'
        ws.cell(row=2, column=4).value = 250
        ws.cell(row=2, column=5).value = 1  # Restaurant ID
        ws.cell(row=2, column=6).value = 'Your Restaurant Name'
        ws.cell(row=2, column=7).value = 1  # Category ID
        ws.cell(row=2, column=8).value = 'Main Course'
        ws.cell(row=2, column=9).value = 'VEG'
        ws.cell(row=2, column=10).value = 'Yes'

        # Instructions sheet
        ws_instructions = wb.create_sheet("Instructions")
        ws_instructions.cell(row=1, column=1).value = "Import Instructions"
        ws_instructions.cell(row=1, column=1).font = Font(bold=True, size=14)

        instructions = [
            "",
            "Required Fields (marked with *):",
            "- Name: Dish name",
            "- Price: Dish price (number)",
            "- Restaurant ID: Existing restaurant ID from your database",
            "",
            "Optional Fields:",
            "- ID: Leave empty for new dishes, fill to update existing",
            "- Description: Dish description",
            "- Category ID: Existing category ID",
            "- Veg Type: VEG, NON_VEG, or EGG",
            "- Is Available: Yes/No (default: Yes)",
            "",
            "Steps to Import:",
            "1. Fill in the dish details in 'Dishes Template' sheet",
            "2. Save the file",
            "3. Run: python manage.py manage_dishes --action import --file yourfile.xlsx",
        ]

        for row_num, instruction in enumerate(instructions, 2):
            ws_instructions.cell(row=row_num, column=1).value = instruction

        # Adjust column widths
        for ws in [wb['Dishes Template'], ws_instructions]:
            for col in ws.columns:
                max_length = 0
                col_letter = col[0].column_letter
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

        wb.save(file_path)
        self.stdout.write(self.style.SUCCESS(f'[OK] Template created: {file_path}'))
